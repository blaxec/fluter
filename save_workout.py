import sys
import json
import os
import openpyxl
from datetime import datetime, timedelta
import re

# Configure standard streams for UTF-8
try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except AttributeError:
    pass # In case of older python or restricted environments

def find_excel_file():
    # Always look relative to the script's own directory first (reliable regardless of cwd)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    script_relative = os.path.join(script_dir, "Новая таблица.xlsx")
    if os.path.exists(script_relative):
        return script_relative

    # Fallback: current working directory
    local_path = "Новая таблица.xlsx"
    if os.path.exists(local_path):
        return local_path
    
    # Fallback: Downloads directory
    downloads_path = os.path.join(os.path.expanduser("~"), "Downloads", "Новая таблица.xlsx")
    if os.path.exists(downloads_path):
        return downloads_path
    
    return None

def main():
    if len(sys.argv) < 2:
        try:
            input_data = sys.stdin.read().strip()
            if not input_data:
                print(json.dumps({"status": "error", "message": "No data provided in arguments or stdin"}))
                return
            data = json.loads(input_data)
        except Exception as e:
            print(json.dumps({"status": "error", "message": f"Failed to read from stdin: {str(e)}"}))
            return
    else:
        try:
            data = json.loads(sys.argv[1])
        except Exception as e:
            print(json.dumps({"status": "error", "message": f"Failed to parse JSON input: {str(e)}"}))
            return

    filepath = find_excel_file()
    if not filepath:
        print(json.dumps({"status": "error", "message": "Excel file 'Новая таблица.xlsx' not found in current folder or Downloads"}))
        return

    day = data.get("day")
    results = data.get("results", [])

    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        
        # Determine column to write to
        # In Row 3, find or create column for today's date
        today = datetime.today()
        today_monday = today - timedelta(days=today.weekday())
        
        col_idx = 6
        target_col = None
        
        while True:
            val = ws.cell(row=3, column=col_idx).value
            if val is None:
                # Found empty column, write today's date
                today_str = today.strftime('%d.%m')
                ws.cell(row=3, column=col_idx, value=today_str)
                target_col = col_idx
                break
            else:
                val_str = str(val).strip()
                # Split existing dates
                parts = re.split(r'[,/;+\s]+', val_str)
                same_week = False
                for p in parts:
                    p = p.strip()
                    if not p:
                        continue
                    try:
                        parsed_d = datetime.strptime(f"{p}.{today.year}", "%d.%m.%Y")
                        parsed_monday = parsed_d - timedelta(days=parsed_d.weekday())
                        if parsed_monday.date() == today_monday.date():
                            same_week = True
                            break
                    except ValueError:
                        pass
                
                if same_week:
                    today_str = today.strftime('%d.%m')
                    if today_str not in parts:
                        new_val = f"{val_str}, {today_str}"
                        ws.cell(row=3, column=col_idx, value=new_val)
                    target_col = col_idx
                    break
            col_idx += 1

        # Define mapping of (day, exercise_name, sub_category, circuit) -> row number
        mapping = {}
        
        # Tuesday (ВТ in Excel)
        for c in [1, 2, 3]:
            mapping[("tuesday", "глубокие отжимания", "", c)] = 5 + c - 1
            mapping[("tuesday", "отжимания уголком", "", c)] = 8 + c - 1
            mapping[("tuesday", "алмазные отжимания", "", c)] = 12 + c - 1
            mapping[("tuesday", "отжимания на возвы-ти", "", c)] = 15 + c - 1
            
        # Thursday (ЧТ in Excel)
        for c in [1, 2, 3]:
            mapping[("thursday", "выпады", "левая нога", c)] = 19 + c - 1
            mapping[("thursday", "выпады", "правая нога", c)] = 22 + c - 1
            mapping[("thursday", "приседания обычные", "", c)] = 25 + c - 1
            
        for c in [1, 2]:
            mapping[("thursday", "подъемы на носках", "левая нога", c)] = 29 + c - 1
            mapping[("thursday", "подъемы на носках", "правая нога", c)] = 31 + c - 1
            mapping[("thursday", "пресс (поднятие ног)", "", c)] = 33 + c - 1
            
        # Friday tab (СБ in Excel)
        # Negatives are on rows 36, 38, 40 for circles 1, 2, 3
        mapping[("friday", "нега-ые подтягивания (сек)", "", 1)] = 36
        mapping[("friday", "нега-ые подтягивания (сек)", "", 2)] = 38
        mapping[("friday", "нега-ые подтягивания (сек)", "", 3)] = 40
        mapping[("friday", "негативные подтягивания", "", 1)] = 36
        mapping[("friday", "негативные подтягивания", "", 2)] = 38
        mapping[("friday", "негативные подтягивания", "", 3)] = 40
        
        for c in [1, 2, 3]:
            mapping[("friday", "тяга гантелей в наклоне", "левая рука", c)] = 42 + c - 1
            mapping[("friday", "тяга гантелей в наклоне", "левая нога", c)] = 42 + c - 1
            mapping[("friday", "тяга гантелей в наклоне", "правая рука", c)] = 45 + c - 1
            mapping[("friday", "тяга гантелей в наклоне", "правая нога", c)] = 45 + c - 1
            mapping[("friday", "подтягивания узким хватом", "", c)] = 48 + c - 1


        written_count = 0
        for r in results:
            ex = r.get("exercise", "").lower().strip()
            sub = r.get("sub_category", "")
            sub_key = sub.lower().strip() if sub else ""
            circuit = int(r.get("circuit", 1))
            val = r.get("value")
            
            key = (day, ex, sub_key, circuit)
            if key in mapping:
                row_num = mapping[key]
                ws.cell(row=row_num, column=target_col, value=val)
                written_count += 1
            else:
                # Try fallback matching without day prefix
                key_no_day = (day, ex, "", circuit)
                if key_no_day in mapping:
                    row_num = mapping[key_no_day]
                    ws.cell(row=row_num, column=target_col, value=val)
                    written_count += 1

        wb.save(filepath)
        print(json.dumps({"status": "success", "filepath": filepath, "column": target_col, "written": written_count}))

    except Exception as e:
        print(json.dumps({"status": "error", "message": f"Error writing to Excel: {str(e)}"}))

if __name__ == '__main__':
    main()
